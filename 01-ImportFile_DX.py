# BSI 面单生成程序，使用中。

import openpyxl
import tkinter as tk
from tkinter import filedialog, messagebox
from datetime import datetime
import os

def DX_ImportFile():
    HEADER_ROW = 4
    DATA_START_ROW = HEADER_ROW + 1
    TARGET_DATA_START = 5

    COPY_FIELDS = ['*客户单号', '*姓名', '*电话', '*国家代码', '*城市', '*邮编', '*地址行1','邮箱']

    CONFIG = {
        'Xhours': {
            'template': 'DX-NEXT-DAY_Model.xlsx',
            'fixed_fields': {
                '*箱号': '01',
                '*长(CM)': 40,
                '*宽(CM)': 30,
                '*高(CM)': 20,
                '*重(KG)': 3,
            }
        },
        '48hours': {
            'template': 'RoyalMail_Model.xlsx',
            'fixed_fields': {
                '*箱号': '01',
                '*长(CM)': 40,
                '*宽(CM)': 30,
                '*高(CM)': 20,
                '*重(KG)': 3,
                '*中文名称': '1111',
                '*英文名': 'aaaa',
            }
        },
        '24hours': {
            'template': 'DPD_Model.xlsx',
            'fixed_fields': {
                '*箱号': '01',
                '*长(CM)': 40,
                '*宽(CM)': 30,
                '*高(CM)': 20,
                '*重(KG)': 3,
                '*中文名称': '1111',
                '*英文名': 'aaaa',
            }
        }
    }

    # 【修改开始】👇👇👇👇👇👇👇👇👇👇👇👇👇👇👇👇👇👇👇👇
    # 判断是否有图形界面环境（Codespaces 没有 DISPLAY）
    if os.environ.get("DISPLAY", "") == "":
        print("⚠️ 当前环境无图形界面（可能在 GitHub Codespaces 中）。")
        print("请输入要处理的 Excel 源文件路径，例如：data/source.xlsx")
        source_file = input("文件路径：").strip()
        if not source_file or not os.path.exists(source_file):
            print(f"❌ 找不到文件：{source_file}")
            return
        print(f"✅ 使用文件：{source_file}")
    else:
        root = tk.Tk()
        root.withdraw()
        source_file = filedialog.askopenfilename(
            title="选择源文件（source.xlsx）",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not source_file:
            messagebox.showwarning("取消操作", "未选择源文件，程序终止。")
            return
    # 【修改结束】👆👆👆👆👆👆👆👆👆👆👆👆👆👆👆👆👆👆👆👆

    source_dir = os.path.dirname(source_file)
    timestamp = datetime.now().strftime("%Y%m%d")  # 只保留日期部分
    log_lines = []

    try:
        src_wb = openpyxl.load_workbook(source_file)
        src_ws = src_wb.active
        src_headers = [cell.value for cell in src_ws[HEADER_ROW]]

        delivery_idx = src_headers.index('时效')
        field_indexes = {
            field: src_headers.index(field) for field in COPY_FIELDS
        }

        targets = {}
        write_rows = {}

        for dtype, conf in CONFIG.items():
            template_path = conf['template']
            if not os.path.exists(template_path):
                # 【修改】GUI 环境下用弹窗；无 GUI 环境下用 print()
                if os.environ.get("DISPLAY", "") == "":
                    print(f"❌ 模板文件不存在：{template_path}")
                else:
                    messagebox.showerror("错误", f"模板文件不存在：{template_path}")
                return

            tgt_wb = openpyxl.load_workbook(template_path)
            tgt_ws = tgt_wb.active
            tgt_headers = [cell.value for cell in tgt_ws[HEADER_ROW]]

            field_indexes_tgt = {
                field: tgt_headers.index(field) for field in COPY_FIELDS if field in tgt_headers
            }

            fill_indexes = {
                key: tgt_headers.index(key) for key in conf['fixed_fields'] if key in tgt_headers
            }

            new_name = f"{os.path.splitext(conf['template'])[0]}_{timestamp}.xlsx"
            new_path = os.path.join(source_dir, new_name)

            targets[dtype] = {
                'workbook': tgt_wb,
                'worksheet': tgt_ws,
                'field_map': field_indexes_tgt,
                'fixed_map': fill_indexes,
                'fixed_values': conf['fixed_fields'],
                'save_path': new_path
            }
            write_rows[dtype] = TARGET_DATA_START

        for i, row in enumerate(src_ws.iter_rows(min_row=DATA_START_ROW, max_col=len(src_headers)), start=DATA_START_ROW):
            raw_delivery = row[delivery_idx].value
            if not raw_delivery:
                log_lines.append(f"第{i}行：时效为空，跳过")
                continue

            delivery = str(raw_delivery).strip()
            if delivery not in targets:
                log_lines.append(f"第{i}行：时效为 [{delivery}]，未匹配配置项，已忽略")
                continue

            order_id_val = row[field_indexes['*客户单号']].value
            if not order_id_val:
                log_lines.append(f"第{i}行：客户单号为空，已忽略")
                continue

            tgt_info = targets[delivery]
            tgt_ws = tgt_info['worksheet']
            write_row = write_rows[delivery]

            for field, src_idx in field_indexes.items():
                if field in tgt_info['field_map']:
                    tgt_col = tgt_info['field_map'][field]
                    tgt_ws.cell(row=write_row, column=tgt_col + 1, value=row[src_idx].value)

            for key, val in tgt_info['fixed_values'].items():
                if key in tgt_info['fixed_map']:
                    tgt_col = tgt_info['fixed_map'][key]
                    tgt_ws.cell(row=write_row, column=tgt_col + 1, value=val)

            write_rows[delivery] += 1
            log_lines.append(f"第{i}行：时效 [{delivery}] 已写入")

        for dtype, info in targets.items():
            info['workbook'].save(info['save_path'])

        log_file = os.path.join(source_dir, f"ImportFile_Log_{timestamp}.txt")
        with open(log_file, "w", encoding="utf-8") as f:
            f.write("\n".join(log_lines))

        msg = "\n".join([
            f"{dtype}：写入 {write_rows[dtype] - TARGET_DATA_START} 行，已保存为\n{targets[dtype]['save_path']}"
            for dtype in write_rows
        ]) + f"\n\n日志已保存：{log_file}"

        # 【修改】根据环境决定用 messagebox 还是 print
        if os.environ.get("DISPLAY", "") == "":
            print("✅ 执行完成：")
            print(msg)
        else:
            messagebox.showinfo("✅ 执行完成", msg)

    except Exception as e:
        # 【修改】根据环境决定如何输出错误
        if os.environ.get("DISPLAY", "") == "":
            print(f"❌ 程序出错：{e}")
        else:
            messagebox.showerror("错误", f"❌ 程序出错：\n{e}")

# 入口
if __name__ == '__main__':
    DX_ImportFile()
